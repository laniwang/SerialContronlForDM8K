import openpyxl
import pprint
# Get all attributes of modules

wb=openpyxl.load_workbook('3rd_control_command.xlsx')
sheet=wb['attribute2']
moduleData={}
a='FLOAT'
b='INT'
c='BOOLEAN'
print(a,b,c)
print('Reading rows')
print(sheet.max_row)
for row in range(4,sheet.max_row+1):
    module=sheet['A'+str(row)].value
    attribute=sheet['B'+str(row)].value
    getState=sheet['C'+str(row)].value
    setState=sheet['D'+str(row)].value
    decState=sheet['E'+str(row)].value
    incState=sheet['F'+str(row)].value
    toggleState=sheet['G'+str(row)].value
    subscribeState=sheet['H'+str(row)].value
    unsubscribeState=sheet['I'+str(row)].value
    datafieldState=sheet['J'+str(row)].value
##    print(sheet['L'+str(row)].value)

    moduleData.setdefault(module,{})
    moduleData[module].setdefault(attribute,{'get':'', 'set':'', 'dec':'', 'inc':'','toggle':'','subscribe':'', 'unsubscribe':'','datafield':'','minValue':'','maxValue':'','dataType':'',})
    moduleData[module][attribute]['get']=getState
    moduleData[module][attribute]['set']=setState
    moduleData[module][attribute]['dec']=decState
    moduleData[module][attribute]['inc']=incState
    moduleData[module][attribute]['toggle']=toggleState
    moduleData[module][attribute]['subscribe']=subscribeState
    moduleData[module][attribute]['unsubscribe']=unsubscribeState
    moduleData[module][attribute]['datafield']=datafieldState
    #new 
    moduleData[module][attribute]['minValue']=sheet['L'+str(row)].value
    moduleData[module][attribute]['maxValue']=sheet['M'+str(row)].value
    moduleData[module][attribute]['dataType']=sheet['N'+str(row)].value
    
    

print('Writing results')
with open ('moduleDm80002.py','a') as f:
    f.write('allmodule = ' +pprint.pformat(moduleData) )
print('Done')
    

      
    

